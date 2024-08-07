import openpyxl

def row_count(file,sheetname):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetname]
    return sheet.max_row
def read_data(file,sheetname,row_no,col_no):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetname]
    return sheet.cell(row=row_no,column=col_no).value
def write_data(file,sheetname,row_no,col_no,data):
    book=openpyxl.load_workbook(file)
    sheet=book[sheetname]
    sheet.cell(row=row_no,column=col_no).value=data
    book.save(file)
